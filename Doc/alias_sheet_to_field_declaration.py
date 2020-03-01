import traceback
import string
import openpyxl
from os import sys, path


GEN_TEMPLATE = '''{
<<items>>
}'''


def get_cell_str(ws, column, row) -> str:
    if column is None or column == '':
        return ''
    cell_item = ws[column + str(row)].value
    if cell_item is None:
        return ''
    return str(cell_item).strip()


def parse(file_name: str,
          field_title: str = 'Field', alias_title: str = 'Alias', comments_title: str = 'Comments',
          padding: int = 8, align: int = 30) -> str:

    wb = openpyxl.load_workbook(file_name)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]

    rows = ws.rows
    columns = ws.columns

    field_column = ''
    alias_column = ''
    comments_column = ''

    for col in string.ascii_uppercase:
        if ws[col + '1'].value == field_title:
            field_column = col
        if ws[col + '1'].value == alias_title:
            alias_column = col
        if ws[col + '1'].value == comments_title:
            comments_column = col

    if field_column == '' or alias_column == '':
        print('Cannot locate field name and alias name column.')
        return ''

    row_index = 1
    gen_items = []
    while True:
        row_index += 1

        field_cell_string = get_cell_str(ws, field_column, row_index)
        alias_cell_string = get_cell_str(ws, alias_column, row_index)
        comments_cell_string = get_cell_str(ws, comments_column, row_index)

        if field_cell_string == '':
            break
        if alias_cell_string == '':
            continue

        middle_padding_count = align - len(field_cell_string)
        middle_padding_count = max(1, middle_padding_count)

        comments_padding_count = align - len(alias_cell_string)
        comments_padding_count = max(1, comments_padding_count)

        prefix_padding = ' ' * padding
        middle_padding = ' ' * middle_padding_count
        comments_padding = ' ' * comments_padding_count
        suffix_comments = '' if comments_cell_string == '' else comments_padding + '# ' + comments_cell_string

        gen_items.append(prefix_padding +
                         "'" + field_cell_string + "'" +
                         ':' + middle_padding +
                         "'" + alias_cell_string + "', " +
                         suffix_comments)
    wb.close()

    if len(gen_items) == 0:
        print('No record can be found..')
        return ''

    return GEN_TEMPLATE.replace('<<items>>', '\n'.join(gen_items))


PARSE_LIST = (
    path.join('AliasSheet', 'Alias_Market_SecuritiesInfo_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Market_IndexInfo_Ts.xlsx'),

    path.join('AliasSheet', 'Alias_Market_TradeCalender_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Market_NamingHistory_Ts.xlsx'),

    path.join('AliasSheet', 'Alias_Finance_Audit_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Finance_BalanceSheet_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Finance_IncomeStatement_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Finance_CashFlowStatement_Ts.xlsx'),

    path.join('AliasSheet', 'Alias_Stockholder_PledgeStatus_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_Stockholder_PledgeHistory_Ts.xlsx'),

    path.join('AliasSheet', 'Alias_TradeData_Index_Daily_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_TradeData_Stock_Daily_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_TradeData_Stock_Daily_AdjFactor_Ts.xlsx'),
    path.join('AliasSheet', 'Alias_TradeData_Stock_Daily_Indicators_Ts.xlsx'),
)


def main():
    for file_name in PARSE_LIST:
        print('-------------------------------' + file_name + '-------------------------------')
        print(parse(file_name))


# ----------------------------------------------------------------------------------------------------------------------

def exception_hook(type, value, tback):
    # log the exception here
    print('Exception hook triggered.')
    print(type)
    print(value)
    print(tback)
    # then call the default handler
    sys.__excepthook__(type, value, tback)


if __name__ == "__main__":
    sys.excepthook = exception_hook
    try:
        main()
    except Exception as e:
        print('Error =>', e)
        print('Error =>', traceback.format_exc())
        exit()
    finally:
        pass
