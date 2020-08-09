import json
from collections import OrderedDict

import openpyxl
import pandas as pd

from .common import *
from .df_utility import *
from .time_utility import *
from ..DataHubEntry import DataHubEntry
from ..Database.DatabaseEntry import DatabaseEntry


def methods_from_prob(prob: dict) -> []:
    return methods_from_method_list(prob.get('methods', []))


def methods_from_method_list(method_list: list) -> []:
    return [method for method, _, _, _ in method_list]


# def standard_dispatch_analysis(securities: [str], methods: [str], data_hub, database,
#                                method_list: list) -> []:
#     result_list = []
#     for query_method in methods:
#         for hash_id, _, _, function_entry in method_list:
#             if hash_id == query_method:
#                 if function_entry is None:
#                     print('Method ' + hash_id + ' not implemented yet.')
#                 else:
#                     try:
#                         result = function_entry(securities, data_hub, database)
#                     except Exception as e:
#                         print('Execute analyzer [' + hash_id + '] Error: ')
#                         print(e)
#                         print(traceback.format_exc())
#                         result = None
#                     finally:
#                         pass
#                     if result is not None and len(result) > 0:
#                         result_list.append((query_method, result))
#                 break
#     return result_list


# ----------------------------------------------------------------------------------------------------------------------

class AnalysisResult:

    SCORE_MIN = 0
    SCORE_MAX = 100
    SCORE_PASS = SCORE_MAX
    SCORE_FAIL = SCORE_MIN
    SCORE_NOT_APPLIED = None

    WEIGHT_NORMAL = 1
    WEIGHT_ONE_VOTE_VETO = 999999

    def __init__(self, securities: str = '', period: datetime.datetime or None = None,
                 score: int or bool = False, reason: str or [str] = '', weight: int = WEIGHT_NORMAL):
        self.method = ''
        self.period = to_py_datetime(period) if period is not None else None
        self.securities = securities

        if isinstance(score, bool):
            self.score = AnalysisResult.SCORE_PASS if score else AnalysisResult.SCORE_FAIL
        elif isinstance(score, (int, float)):
            self.score = score
            self.score = min(self.score, AnalysisResult.SCORE_MAX)
            self.score = max(self.score, AnalysisResult.SCORE_MIN)
        else:
            self.score = score

        if reason is None:
            self.reason = ''
        elif isinstance(reason, (list, tuple)):
            self.reason = '\n'.join(reason)
        else:
            self.reason = reason

        self.weight = weight

    def pack(self, to_str: bool = True) -> dict:
        return {
            # Make the dict keys 'happens to' the fields of Result.Analyzer
            'period': str(self.period) if to_str else self.period,
            'analyzer': str(self.method) if to_str else self.method,
            'stock_identity': str(self.securities) if to_str else self.securities,
            
            'score': str(self.score) if to_str else self.score,
            'reason': str(self.reason) if to_str else self.reason,
            'weight': str(self.weight) if to_str else self.weight,
        }

    def unpack(self, data: dict):
        period = data.get('period', 'None')
        self.period = None if period == 'None' else text_auto_time(period)

        self.method = data.get('analyzer', '')
        self.securities = data.get('stock_identity', '')

        score = data.get('score', AnalysisResult.SCORE_NOT_APPLIED)
        self.score = str2float_safe(score, AnalysisResult.SCORE_NOT_APPLIED)

        self.reason = data.get('reason', [])
        self.weight = float(data.get('weight', AnalysisResult.WEIGHT_NORMAL))

    def serialize(self) -> str:
        return json.dumps(self.pack())
    
    def deserialize(self, json_text: str):
        self.unpack(json.loads(json_text))

    def rough_size(self) -> int:
        return sys.getsizeof(self.period) + \
               sys.getsizeof(self.method) + \
               sys.getsizeof(self.securities) + \
               sys.getsizeof(self.score) +  \
               sys.getsizeof(self.reason) +  \
               sys.getsizeof(self.weight)


# --------------------------------------------- Analysis Result Conversion ---------------------------------------------

def analysis_results_to_json(result_list: [AnalysisResult], fp=None) -> bool or str:
    def _analysis_result_json_hook(analysis_result: AnalysisResult) ->dict:
        if isinstance(analysis_result, AnalysisResult):
            return analysis_result.pack(True)
        else:
            print('Unknown class: ' + str(analysis_result))
            return {}
    return json.dump(result_list, fp, default=_analysis_result_json_hook, sort_keys=True, indent=4) \
        if fp is not None else json.dumps(result_list, default=_analysis_result_json_hook, sort_keys=True, indent=4)


def analysis_results_from_json(fp) -> [AnalysisResult]:
    def _json_analysis_result_hook(_dict: dict) -> AnalysisResult:
        analysis_result = AnalysisResult()
        analysis_result.unpack(_dict)
        return analysis_result
    if isinstance(fp, str):
        return json.loads(fp, object_hook=_json_analysis_result_hook)
    else:
        return json.load(fp, object_hook=_json_analysis_result_hook)


def analysis_result_list_to_table(result_list: [AnalysisResult]) -> {str: {str: [AnalysisResult]}}:
    result_table = OrderedDict()
    for analysis_result in result_list:
        analyzer_uuid = analysis_result.method
        stock_identity = analysis_result.securities
        if analyzer_uuid not in result_table.keys():
            result_table[analyzer_uuid] = OrderedDict()
        if stock_identity not in result_table[analyzer_uuid].keys():
            result_table[analyzer_uuid][stock_identity] = []
        result_table[analyzer_uuid][stock_identity].append(analysis_result)
    return result_table


def get_security_result_from_analysis_result_list(result_list: [AnalysisResult],
                                                  stock_identity: str) -> [AnalysisResult]:
    security_result_list = []
    for analysis_result in result_list:
        identity = analysis_result.securities
        if identity == stock_identity:
            security_result_list.append(analysis_result)
    return security_result_list


def analysis_result_list_to_single_stock_report(result_list: [AnalysisResult], stock_ideneity: str) -> pd.DataFrame:
    security_result_list = get_security_result_from_analysis_result_list(result_list, stock_ideneity)
    result_table = {}
    for analysis_result in security_result_list:
        analyzer_uuid = analysis_result.method
        if analyzer_uuid not in result_table.keys():
            result_table[analyzer_uuid] = [analysis_result]
        else:
            result_table[analyzer_uuid].append(analysis_result)
    result_report = None

    # TODO: Check Duplicate: {"period" : ISODate("2015-12-31T00:00:00Z"),"stock_identity" : "600103.SSE", "method":"7e132f82-a28e-4aa9-aaa6-81fa3692b10c"}
    for analyzer_uuid, result_list in result_table.items():
        # content = [r.reason + ' | ' + str(r.score) for r in result_list]
        # indexes = [r.period for r in result_list]

        content = []
        indexes = []
        for r in result_list:
            text = r.reason if str_available(r.reason) else 'OK'
            text += ' [' + str(r.score) + ']'
            content.append(text)
            indexes.append(r.period)

        s = pd.Series(content, index=indexes)
        s.name = analyzer_uuid

        df = s.to_frame()
        df = df.groupby(df.index).first()
        result_report = df if result_report is None else pd.concat([result_report, df], axis=1)
    return result_report.sort_index(ascending=False)


def analysis_dataframe_to_list(df: pd.DataFrame) -> [AnalysisResult]:
    if df is None or df.empty:
        return []

    result_list = []
    # data_dict = df.T.apply(lambda x: x.dropna().to_dict()).tolist()
    for period, analyzer, stock_identity, score, reason, weight in \
            zip(df['period'], df['analyzer'], df['stock_identity'], df['score'], df['reason'], df['weight']):
        analysis_result = AnalysisResult()

        analysis_result.period = to_py_datetime(period) if period is not None else None
        analysis_result.method = analyzer
        analysis_result.securities = stock_identity

        analysis_result.score = str2float_safe(score, None)
        analysis_result.reason = reason
        analysis_result.weight = str2float_safe(weight, None)

        if str_available(analyzer):
            result_list.append(analysis_result)
    return result_list


# ----------------------------------------------------------------------------------------------------------------------

class AnalysisContext:
    def __init__(self):
        self.cache = {}
        self.extra = {}
        self.logger = None
        self.progress = None


# ----------------------------------------------------------------------------------------------------------------------

def function_entry_example(securities: str, time_serial: tuple, data_hub: DataHubEntry,
                           database: DatabaseEntry, context: AnalysisContext, **kwargs) -> AnalysisResult:
    """
    The example of analyzer function entry.
    :param securities: A single securities code, should be a str.
    :param time_serial: The analysis period
    :param data_hub:  DataHubEntry type
    :param database: DatabaseEntry type
    :param context: AnalysisContext type, which can hold cache data for multiple analysis
    :return: AnalysisResult
    """
    pass


method_list_example = [
    ('5c496d06-9961-4157-8d3e-a90683d6d32c', 'analyzer brief', 'analyzer details', function_entry_example),
]


# ----------------------------------------------------------------------------------------------------------------------

def standard_dispatch_analysis(methods: [str], securities: [str], time_serial: tuple,
                               data_hub, database, extra: dict, method_list: list) -> [(str, [])] or None:
    context = AnalysisContext()
    if isinstance(extra, dict):
        context.extra = extra
        context.sas = extra.get('sas', None)
        context.logger = extra.get('logger', print)
        context.progress = extra.get('progress', ProgressRate())

    result_list = []
    for query_method in methods:
        sub_list = []
        context.cache.clear()
        for _uuid, _, _, function_entry in method_list:
            if _uuid != query_method:
                continue
            if function_entry is None:
                print('Method ' + _uuid + ' not implemented yet.')
                break
            context.progress.set_progress(_uuid, 0, len(securities))
            for s in securities:
                try:
                    result = function_entry(s, time_serial, data_hub, database, context, **extra)
                except Exception as e:
                    error_info = 'Execute analyzer [' + _uuid + '] for [' + s + '] got exception.'
                    print(error_info)
                    print(e)
                    print(traceback.format_exc())
                    result = AnalysisResult(s, None, AnalysisResult.SCORE_NOT_APPLIED, error_info)
                finally:
                    context.progress.increase_progress(_uuid)
                    # print('Analyzer %s progress: %.2f%%' % (_uuid, context.progress.get_progress_rate(_uuid) * 100))
                if result is None:
                    result = AnalysisResult(s, None, AnalysisResult.SCORE_NOT_APPLIED, 'NONE')
                if not isinstance(result, (list, tuple)):
                    result = [result]
                for r in result:
                    r.method = _uuid
                sub_list.extend(result)

            # # Fill result list for alignment
            # while len(sub_list) < len(securities):
            #     sub_list.append([AnalysisResult(securities[len(sub_list)],
            #     None, AnalysisResult.SCORE_NOT_APPLIED, 'NONE')])

            result_list.extend(sub_list)
            context.progress.set_progress(_uuid, len(securities), len(securities))
    return result_list if len(result_list) > 0 else None


# --------------------------------------------------- Analyzer Helper --------------------------------------------------

# def check_append_report_when_data_missing(df: pd.DataFrame, securities: str,
#                                           uri: str, fields: str or [str], result: list):
#     if df is None or len(df) == 0:
#         error_info = uri + ': Cannot find data for securities : ' + securities
#         log_error(error_info)
#         result.append(AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info))
#         return True
#     if not isinstance(fields, (list, tuple)):
#         fields = [fields]
#     for field in fields:
#         if field not in df.columns:
#             error_info = uri + ': Field ' + field + ' missing for securities : ' + securities
#             log_error(error_info)
#             result.append(AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info))
#             return True
#     return False


def check_gen_report_when_data_missing(df: pd.DataFrame, securities: str,
                                       uri: str, fields: str or [str]) -> AnalysisResult or None:
    if df is None or len(df) == 0:
        error_info = uri + ': Cannot find data for securities : ' + securities
        log_error(error_info)
        return AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED, error_info)
    if not isinstance(fields, (list, tuple)):
        fields = [fields]
    for field in fields:
        if field not in df.columns:
            error_info = uri + ': Field ' + field + ' missing for securities : ' + securities
            log_error(error_info)
            return AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED, error_info)
    return None


def gen_report_when_analyzing_error(securities: str, exception: Exception):
    error_info = 'Error when analysing  : ' + securities + '\n'
    error_info += str(exception)
    log_error(error_info)
    print(traceback.format_exc())
    return AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED, error_info)


def batch_query_readable_annual_report_pattern(
        data_hub, securities: str, time_serial: tuple,
        fields_balance_sheet: [str] = None,
        fields_income_statement: [str] = None,
        fields_cash_flow_statement: [str] = None) -> (pd.DataFrame, AnalysisResult):

    df = None
    if fields_balance_sheet is not None and len(fields_balance_sheet) > 0:
        df_balance, result = query_readable_annual_report_pattern(
            data_hub, 'Finance.BalanceSheet', securities, time_serial, fields_balance_sheet)
        if result is not None:
            return df, result
        df = df_balance

    if fields_income_statement is not None and len(fields_income_statement) > 0:
        df_income, result = query_readable_annual_report_pattern(
            data_hub, 'Finance.IncomeStatement', securities, time_serial, fields_income_statement)
        if result is not None:
            return df, result
        df = df_income if df is None else (pd.merge(df, df_income, how='left', on=['stock_identity', 'period']))

    if fields_cash_flow_statement is not None and len(fields_cash_flow_statement) > 0:
        df_cash, result = query_readable_annual_report_pattern(
            data_hub, 'Finance.CashFlowStatement', securities, time_serial, fields_cash_flow_statement)
        if result is not None:
            return df, result
        df = df_cash if df is None else (pd.merge(df, df_cash, how='left', on=['stock_identity', 'period']))
    df = df.sort_values('period')
    return df, None


def query_readable_annual_report_pattern(data_hub, uri: str, securities: str, time_serial: tuple,
                                         fields: [str]) -> (pd.DataFrame, AnalysisResult):
    """
    The pattern of query readable annual report. It will do the following things:
    1. Check readable names are all known
    2. Query data from data center
    3. Only keep annual report
    4. Check empty
    5. Fill na with 0.0 and sort by date
    :param data_hub: The instance of DataHubEntry
    :param uri: The uri user queries for
    :param securities: The securities user queries
    :param time_serial: The data range user queries
    :param fields: The readable fields user queries
    :return: (Query Result if successful, else None, Analysis Result if fail else None)
    """
    if not data_hub.get_data_center().check_readable_name(fields):
        return None, AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED, 'Unknown readable name detect.')

    fields_stripped = list(set(fields + ['stock_identity', 'period']))
    df = data_hub.get_data_center().query(uri, securities, time_serial, fields=fields_stripped, readable=True)
    if df is None or len(df) == 0:
        return None, AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED,
                                    'No data, skipped' + str(time_serial))

    # Only analysis annual report
    df = df[df['period'].dt.month == 12]

    if len(df) == 0:
        return None, AnalysisResult(securities, None, AnalysisResult.SCORE_NOT_APPLIED,
                                    'No data in this period' + str(time_serial))
    df.fillna(0.0, inplace=True)
    df = df.sort_values('period', ascending=False)

    return df, None


def check_industry_in(securities: str, industries: [str], data_hub: DataHubEntry,
                      database: DatabaseEntry, context: AnalysisContext) -> bool:
    nop(database)

    if context.cache.get('securities_info', None) is None:
        context.cache['securities_info'] = data_hub.get_data_center().query('Market.SecuritiesInfo')
    df_info = context.cache.get('securities_info', None)

    df_slice = df_info[df_info['stock_identity'] == securities]
    industry = get_dataframe_slice_item(df_slice, 'industry', 0, '')

    return industry in industries


# ---------------------------------------------------- Parse Result ----------------------------------------------------

"""
The results should look like:

method1           method2           method3           ...           methodM
m1_result1        m2_result1        m3_result1                      mM_result1
m1_result2        m2_result2        m3_result2                      mM_result2
m1_result3        m2_result3        m3_result3                      mM_result3
.                 .                 .                               .
.                 .                 .                               .
.                 .                 .                               .
m1_resultN        m2_resultN        m3_resultN                      mM_resultN
"""


def get_securities_in_result(result: dict) -> [str]:
    securities = []
    for method, results in result.items():
        for r in results:
            if str_available(r.securities) and r.securities not in securities:
                securities.append(r.securities)
    return securities


def pick_up_pass_securities(result: dict, score_threshold: int, not_applied_as_fail: bool = False) -> [str]:
    securities = get_securities_in_result(result)
    for method, results in result.items():
        for r in results:
            if r.score == AnalysisResult.SCORE_NOT_APPLIED:
                exclude = not_applied_as_fail
            else:
                exclude = (r.score < score_threshold)
            if exclude and r.securities in securities:
                securities.remove(r.securities)
    return securities


# ---------------------------------------------------- Excel Report ----------------------------------------------------


fill_pass = openpyxl.styles.PatternFill(patternType="solid", start_color="10EF10")
fill_flaw = openpyxl.styles.PatternFill(patternType="solid", start_color="FFFF10")
fill_fail = openpyxl.styles.PatternFill(patternType="solid", start_color="EF1010")
fill_none = openpyxl.styles.PatternFill(patternType="solid", start_color="808080")


def __score_to_fill_style(score: int or None):
    if score is None:
        fill_style = fill_none
    elif score < 50:
        fill_style = fill_fail
    elif score <= 75:
        fill_style = fill_flaw
    else:
        fill_style = fill_pass
    return fill_style


def __score_to_fill_text(score: int or None):
    if score is None:
        return '-'
    elif score == 0:
        return 'VETO'
    elif score <= 50:
        return 'FAIL'
    elif score <= 75:
        return 'FLAW'
    elif score <= 90:
        return 'WELL'
    else:
        return 'PASS'


def __aggregate_single_security_results(results: [AnalysisResult]) -> (int, int, str):
    veto = False
    score = []
    reason = []
    weight = AnalysisResult.WEIGHT_NORMAL
    for r in results:
        if r.score is not None:
            if r.weight == AnalysisResult.WEIGHT_ONE_VOTE_VETO and r.score == AnalysisResult.SCORE_FAIL:
                veto = True
            if isinstance(r.score, (int, float)):
                score.append(r.score)
            else:
                print('Error score.')
        if str_available(r.reason):
            reason.append(r.reason)
        if r.weight is not None:
            weight = max(weight, r.weight)
    if veto:
        score = AnalysisResult.SCORE_FAIL
    elif len(score) == 0:
        score = None
    else:
        score = int(sum(score) / len(score))
    reason = '\n'.join(reason)
    return score, weight, reason


def __calc_avg_score_with_weight(scores: list, weights: list) -> int or None:
    sum_score = 0
    sum_weight = 0
    for score, weight in zip(scores, weights):
        if score is None:
            continue
        elif weight == AnalysisResult.WEIGHT_ONE_VOTE_VETO:
            if score != AnalysisResult.SCORE_PASS:
                return 0
        else:
            sum_score += score * weight
            sum_weight += weight
    return (sum_score / sum_weight) if sum_weight > 0 else None


def generate_analysis_report(result: dict, file_path: str, analyzer_name_dict: dict = {}, stock_name_dict: dict = {},
                             extra_data: pd.DataFrame = None):
    """
    Format of result: {analyzer_uuid: {security_identity:［AnalysisResult］}}
    """
    if result is None or len(result) == 0:
        return

    wb = openpyxl.Workbook()
    ws_score = wb.active
    ws_score.title = 'Score'
    ws_comments = wb.create_sheet('Comments')

    ws_score['A1'] = 'Securities\\Analyzer'
    ws_comments['A1'] = 'Securities\\Analyzer'

    ROW_OFFSET = 2
    all_weight = []
    all_score = []

    column = 1
    for analyzer_uuid, analysis_result in result.items():
        # Note that this function will generate the report column by column
        #   The first column is the securities code and name
        #   The first row of each column is the name of analyzer
        # So we need to record the score for each cell, then we can calculate the total score by row at the end

        if len(all_score) < len(analysis_result):
            print('%s : Result buffer increased: %d -> %d' % (analyzer_uuid, len(all_score), len(analysis_result)))
            while len(all_score) < len(analysis_result):
                all_score.append([])
                all_weight.append([])

        # Write securities column
        if column == 1:
            # The first run. Init the total score list here.
            # Flaw: The first column of result should be the full one. Otherwise the index may out of range.
            # all_score = [[] for _ in range(0, len(analysis_result))]
            # all_weight = [[] for _ in range(0, len(analysis_result))]
            #
            # if len(all_score) != len(analysis_result) or len(all_weight) != len(analysis_result):
            #     print('Error: list length not as expect.')
            #     assert False

            # Collect and sort securities list as the following order
            securities_list = sorted(analysis_result.keys())

            row = 2
            col = index_to_excel_column_name(column)
            for security, results in analysis_result.items():
                securities_name = stock_name_dict.get(security, '')
                display_text = (security + ' | ' + securities_name) if securities_name != '' else security
                ws_score[col + str(row)] = display_text
                ws_comments[col + str(row)] = display_text
                row += 1
            column = 2

        # Write analyzer name
        row = 1
        col = index_to_excel_column_name(column)
        analyzer_name = analyzer_name_dict.get(analyzer_uuid, analyzer_uuid)
        ws_score[col + str(row)] = analyzer_name
        ws_comments[col + str(row)] = analyzer_name

        # Write scores
        row = ROW_OFFSET
        for security in securities_list:
            results = analysis_result.get(security, None)
            if results is not None:
                score, weight, reason = __aggregate_single_security_results(results)
            else:
                score, weight, reason = AnalysisResult.SCORE_NOT_APPLIED, 0, 'x'

            ws_score[col + str(row)] = score
            ws_comments[col + str(row)] = reason

            # DEBUG: Catch issue
            try:
                if score is not None:
                    all_score[row - ROW_OFFSET].append(score)
                    all_weight[row - ROW_OFFSET].append(weight)
                fill_style = __score_to_fill_style(score)

                ws_score[col + str(row)].fill = fill_style
                ws_comments[col + str(row)].fill = fill_style
            except Exception as e:
                print(e)
                print('Catch')
            finally:
                pass

            row += 1
        column += 1

    # Write total score
    row = 1
    col_rate = index_to_excel_column_name(column)
    col_vote = index_to_excel_column_name(column + 1)

    for scores, weights in zip(all_score, all_weight):
        if row == 1:
            ws_score[col_vote + str(row)] = 'Vote'
            ws_comments[col_vote + str(row)] = 'Vote'
            ws_score[col_rate + str(row)] = 'Total Rate'
            ws_comments[col_rate + str(row)] = 'Total Rate'
            row = 2

        if len(scores) > 0:
            # min_score = min(score)
            avg_score = __calc_avg_score_with_weight(scores, weights)
        else:
            # min_score = None
            avg_score = None
        # fill_text = __score_to_fill_text(min_score)
        # fill_style = __score_to_fill_style(min_score)

        fill_text = __score_to_fill_text(avg_score)
        fill_style = __score_to_fill_style(avg_score)

        # ------------------- Rate -------------------

        if avg_score is not None:
            ws_score[col_rate + str(row)] = str(int(avg_score))
            ws_comments[col_rate + str(row)] = str(int(avg_score))
        else:
            ws_score[col_rate + str(row)] = '-'
            ws_comments[col_rate + str(row)] = '-'

        ws_score[col_rate + str(row)].fill = fill_style
        ws_comments[col_rate + str(row)].fill = fill_style

        # ------------------- Vote -------------------

        ws_score[col_vote + str(row)] = fill_text
        ws_comments[col_vote + str(row)] = fill_text

        ws_score[col_vote + str(row)].fill = fill_style
        ws_comments[col_vote + str(row)].fill = fill_style

        row += 1

    # Write the extra data

    if isinstance(extra_data, pd.DataFrame):
        row = 1
        column += 2
        col_extra = index_to_excel_column_name(column)

        ws_score[col_extra + str(row)] = 'Extra Data: '
        ws_comments[col_extra + str(row)] = 'Extra Data: '
        column += 1

        alignment_df = pd.DataFrame({'stock_identity': securities_list})
        merged_df = pd.merge(alignment_df, extra_data, how='left', on='stock_identity')
        merged_df = merged_df.fillna('-')
        del merged_df['stock_identity']

        columns = merged_df.columns
        for title in columns:
            row = 1
            col_extra = index_to_excel_column_name(column)
            ws_score[col_extra + str(row)] = title
            ws_comments[col_extra + str(row)] = title

            row = 2
            column_data = merged_df[title]
            for serial_item in column_data:
                ws_score[col_extra + str(row)] = serial_item
                ws_comments[col_extra + str(row)] = serial_item
                row += 1
            column += 1

    # Write file
    wb.save(file_path)


# --------------------------------------------------------- UI ---------------------------------------------------------

from StockAnalysisSystem.core.Utiltity.ui_utility import *
from StockAnalysisSystem.core.Utiltity.TableViewEx import *


class AnalyzerSelector(QDialog):
    TABLE_HEADER_ANALYZER = ['', 'Strategy', 'Comments', 'UUID']

    def __init__(self, analyzer_utility):
        super(AnalyzerSelector, self).__init__()
        self.__analyzer_utility = analyzer_utility
        self.__ok = True
        self.__table_analyzer = TableViewEx()
        self.__button_ok = QPushButton('OK')
        self.__button_cancel = QPushButton('Cancel')
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)

        layout.addWidget(self.__table_analyzer)
        layout.addLayout(horizon_layout([QLabel(''), self.__button_ok, self.__button_cancel], [8, 1, 1]))

        self.__table_analyzer.SetCheckableColumn(0)
        self.__table_analyzer.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        self.__button_ok.clicked.connect(self.__on_button_ok)
        self.__button_cancel.clicked.connect(self.__on_button_cancel)

        self.setMinimumSize(800, 600)
        self.setWindowTitle('Select Analyzer')

        self.load_analyzer()

    def __on_button_ok(self):
        self.__ok = True
        self.close()

    def __on_button_cancel(self):
        self.close()

    def is_ok(self) -> bool:
        return self.__ok

    def load_analyzer(self):
        self.__table_analyzer.Clear()
        self.__table_analyzer.SetRowCount(0)
        self.__table_analyzer.SetColumn(AnalyzerSelector.TABLE_HEADER_ANALYZER)

        analyzer_info = self.__analyzer_utility.analyzer_info()
        for analyzer_uuid, analyzer_name, analyzer_detail, _ in analyzer_info:
            self.__table_analyzer.AppendRow(['', analyzer_name, analyzer_detail, analyzer_uuid])

    def get_select_strategy(self):
        analyzer_list = []
        for i in range(self.__table_analyzer.RowCount()):
            if self.__table_analyzer.GetItemCheckState(i, 0) == QtCore.Qt.Checked:
                uuid = self.__table_analyzer.GetItemText(i, 3)
                analyzer_list.append(uuid)
        return analyzer_list




