import logging
import traceback

from os import sys, path

import openpyxl
from PyQt5.QtGui import QFontMetrics
from PyQt5.QtWidgets import QLineEdit, QCheckBox, QWidget, QComboBox, QDateTimeEdit, QFileDialog, QRadioButton

from StockAnalysisSystem.core.Utility.common import *
from StockAnalysisSystem.core.Utility.ui_utility import *
from StockAnalysisSystem.core.Utility.time_utility import *
from StockAnalysisSystem.ui.Utility.ui_context import UiContext
from StockAnalysisSystem.interface.interface import SasInterface as sasIF
from StockAnalysisSystem.core.Utility.TableViewEx import TableViewEx
from StockAnalysisSystem.core.Utility.securities_selector import SecuritiesPicker


# ------------------------------------------------------ ExportUi ------------------------------------------------------

class ExportUi(QDialog):
    TABLE_HEADER = ['', 'Field', 'Group By', 'Align By']
    INDEX_SELECT_CHECK = 0
    INDEX_GROUP_CHECK = 2
    INDEX_ALIGN_CHECK = 3

    FILL_GROUP = openpyxl.styles.PatternFill(patternType="solid", start_color="EE1111")

    def __init__(self, export_data: pd.DataFrame):
        super(ExportUi, self).__init__()
        self.__export_data = export_data

        self.__radio_vertical = QRadioButton('Vertical')
        self.__radio_horizon = QRadioButton('Horizon')
        self.__check_readable = QCheckBox('Readable')
        self.__button_export = QPushButton('Export')
        self.__table_field_config = TableViewEx()

        self.init_ui()
        self.update_table()

    def init_ui(self):
        self.__layout_control()
        self.__config_control()

    def __layout_control(self):
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        main_layout.addWidget(self.__table_field_config)

        line = QHBoxLayout()
        line.addStretch()
        line.addWidget(self.__radio_vertical)
        line.addWidget(self.__radio_horizon)
        line.addWidget(self.__check_readable)
        line.addWidget(self.__button_export)
        main_layout.addLayout(line)

    def __config_control(self):
        self.__radio_horizon.setChecked(True)
        self.__check_readable.setChecked(True)

        self.__button_export.clicked.connect(self.on_button_export)

        self.__table_field_config.SetColumn(ExportUi.TABLE_HEADER)
        self.__table_field_config.SetCheckableColumn(ExportUi.INDEX_SELECT_CHECK)
        self.__table_field_config.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        self.setMinimumSize(QSize(350, 600))
        self.setWindowTitle('Export Option')

    def on_group_clicked(self, index: int):
        check_group = self.__table_field_config.GetItemData(index, ExportUi.INDEX_GROUP_CHECK)
        if check_group.isChecked():
            self.__table_field_config.SetItemCheckState(index, 0, Qt.Checked)

    def on_align_clicked(self, index: int):
        check_align = self.__table_field_config.GetItemData(index, ExportUi.INDEX_ALIGN_CHECK)
        if check_align.isChecked():
            self.__table_field_config.SetItemCheckState(index, 0, Qt.Checked)
            for i in range(self.__table_field_config.RowCount()):
                if i != index:
                    check_align = self.__table_field_config.GetItemData(i, ExportUi.INDEX_ALIGN_CHECK)
                    check_align.setChecked(False)

    def on_button_export(self):
        align_by = ''
        group_by = []
        select_name = []
        for i in range(self.__table_field_config.RowCount()):
            field_name = self.__table_field_config.GetItemData(i, 1)
            check_group = self.__table_field_config.GetItemData(i, ExportUi.INDEX_GROUP_CHECK)
            check_align = self.__table_field_config.GetItemData(i, ExportUi.INDEX_ALIGN_CHECK)
            selected_field = False
            if check_group.isChecked():
                selected_field = True
                group_by.append(field_name)
            if check_align.isChecked():
                selected_field = True
                align_by = field_name
            if self.__table_field_config.GetItemCheckState(i, 0) == Qt.Checked:
                selected_field = True
            if selected_field:
                select_name.append(field_name)

        if len(select_name) == 0:
            QMessageBox().information(self, 'Warning.', 'Please select export field.')
            return

        file_path, ok = QFileDialog().getSaveFileName(self, 'Select Export Excel Path', '',
                                                      'XLSX Files (*.xlsx);;All Files (*)')
        if ok:
            self.export_excel(file_path, self.__export_data, select_name, group_by, align_by,
                              self.__radio_vertical.isChecked(), self.__check_readable.isChecked())
            QMessageBox().information(self, 'Information.', 'Export finished. File saved at:\n  %s' % file_path)
            self.close()

    def update_table(self):
        columns = list(self.__export_data.columns)
        self.__table_field_config.Clear()
        self.__table_field_config.SetColumn(ExportUi.TABLE_HEADER)

        for field_name in columns:
            self.__table_field_config.AppendRow(['', field_name, '', ''])
            index = self.__table_field_config.RowCount() - 1

            check_group = QCheckBox('')
            check_group.clicked.connect(partial(self.on_group_clicked, index))
            self.__table_field_config.SetItemData(index, ExportUi.INDEX_GROUP_CHECK, check_group)
            self.__table_field_config.SetCellWidget(index, ExportUi.INDEX_GROUP_CHECK, check_group)

            check_align = QCheckBox('')
            check_align.clicked.connect(partial(self.on_align_clicked, index))
            self.__table_field_config.SetItemData(index, ExportUi.INDEX_ALIGN_CHECK, check_align)
            self.__table_field_config.SetCellWidget(index, ExportUi.INDEX_ALIGN_CHECK, check_align)

            self.__table_field_config.SetItemData(index, 1, field_name)
            self.__table_field_config.SetItemCheckState(index, 0, Qt.Checked)

    def export_excel(self, file_path: str, df: pd.DataFrame, fields: [str], group_by: [str],
                     align_by: str, vertical: bool = True, readable: bool = False):
        export_data = df[fields]
        if readable:
            # TODO: Add readable interface and make field radable
            pass

        wb = openpyxl.Workbook()
        ws = wb.active

        if len(group_by) > 0:
            align_column = self.__prepare_align_columns(export_data, align_by)
            
            write_row, write_col = 1, 1
            grouped_df = export_data.groupby(group_by)
            for group, df in grouped_df:
                if align_column is not None:
                    df = pd.merge(align_column, df, on=align_by, how='left')

                cell = excel_cell(write_row, write_col)
                ws[cell] = str(group)
                ws[cell].fill = ExportUi.FILL_GROUP
                if vertical:
                    _, write_col = self.__write_df_to_excel_vertical(ws, df, 1, write_col + 1)
                else:
                    write_row, _ = self.__write_df_to_excel_horizon(ws, df, write_row + 1, 1)
        else:
            if vertical:
                self.__write_df_to_excel_vertical(ws, df, 1, 1)
            else:
                self.__write_df_to_excel_horizon(ws, df, 1, 1)

        wb.save(file_path)

    def __prepare_align_columns(self, df: pd.DataFrame, align_by: str):
        if str_available(align_by):
            align_column = df[align_by]
            align_column = align_column.drop_duplicates()
            align_column = align_column.sort_values()
            return align_column
        else:
            return None

    def __write_df_to_excel_vertical(self, ws, df: pd.DataFrame, start_row: int, start_col: int) -> (int, int):
        write_row = start_row
        write_col = start_col
        for field in df.columns:
            cell = excel_cell(write_row, write_col)
            ws[cell] = field
            write_col += 1

        write_col = start_col
        for field in df.columns:
            write_row = start_row + 1
            data_column = df[field]
            for item in data_column:
                cell = excel_cell(write_row, write_col)
                ws[cell] = str(item)
                write_row += 1
            write_col += 1
        return write_row, write_col

    def __write_df_to_excel_horizon(self, ws, df: pd.DataFrame, start_row: int, start_col: int) -> (int, int):
        write_row = start_row
        write_col = start_col
        for field in df.columns:
            cell = excel_cell(write_row, write_col)
            ws[cell] = field
            write_row += 1

        write_row = start_row
        for field in df.columns:
            write_col = start_col + 1
            data_column = df[field]
            for item in data_column:
                cell = excel_cell(write_row, write_col)
                ws[cell] = str(item)
                write_col += 1
            write_row += 1
        return write_row, write_col


# ---------------------------------------------------- QueryOption -----------------------------------------------------

class QueryOption(QDialog):
    def __init__(self, tip_text: str, selection: int = 0, intersect_columns: [str] = list(), force_selection: bool = False):
        super(QueryOption, self).__init__()

        self.__is_ok = False
        self.__tip_text = tip_text
        self.__selection = selection
        self.__force_selection = force_selection
        self.__intersect_columns = intersect_columns

        self.__radio_replace = QRadioButton('Replace - Clear and show new data')
        self.__radio_concat = QRadioButton('Concat - Extend row, column alignment by name')
        self.__radio_merge = QRadioButton('Merge - Extend column, row alignment by specified field (reserved)')

        self.__button_ok = QPushButton('OK')
        self.__button_cancel = QPushButton('Cancel')
        self.__table_merge_on = TableViewEx()

        self.init_ui()

    def init_ui(self):
        root_layout = QVBoxLayout()
        self.setLayout(root_layout)

        root_layout.addWidget(QLabel(self.__tip_text))
        root_layout.addWidget(self.__radio_replace)
        root_layout.addWidget(self.__radio_concat)
        root_layout.addWidget(self.__radio_merge)

        line = QHBoxLayout()
        line.addStretch()
        line.addWidget(self.__button_ok)
        line.addWidget(self.__button_cancel)
        root_layout.addLayout(line)

        if self.__selection == 0:
            self.__radio_replace.setChecked(True)
        elif self.__selection == 1:
            self.__radio_concat.setChecked(True)
        else:
            self.__radio_replace.setChecked(True)

        self.__radio_replace.setEnabled(not self.__force_selection)
        self.__radio_concat.setEnabled(not self.__force_selection)
        self.__radio_merge.setEnabled(False)

        self.__button_ok.clicked.connect(self.on_button_ok)
        self.__button_cancel.clicked.connect(self.on_button_cancel)

        self.setWindowTitle('Result Option')

    def on_button_ok(self):
        if self.__radio_replace.isChecked():
            self.__selection = 0
        elif self.__radio_concat.isChecked():
            self.__selection = 1
        elif self.__radio_merge.isChecked():
            self.__selection = 2
        else:
            self.__selection = -1
        self.__is_ok = True
        self.close()

    def on_button_cancel(self):
        self.__is_ok = False
        self.close()

    def is_ok(self) -> bool:
        return self.__is_ok

    def get_selection(self) -> [str]:
        return self.__selection


# ----------------------------------------------------- DataHubUi ------------------------------------------------------

class DataHubUi(QWidget):
    def __init__(self, context: UiContext):
        super(DataHubUi, self).__init__()

        self.__context = context
        self.__translate = QtCore.QCoreApplication.translate
        self.__select_list = []
        self.__query_result = pd.DataFrame()

        self.__combo_uri = QComboBox()

        self.__text_selected = QTextEdit('')
        self.__button_reset = QPushButton(self.__translate('', 'Reset'))
        self.__button_select = QPushButton(self.__translate('', 'Select'))

        # self.__line_identity = QLineEdit()
        # if self.__context is not None:
        #     self.__combo_identity = SecuritiesSelector(context.get_sas_interface())
        # else:
        #     self.__combo_identity = QComboBox()

        self.__table_main = EasyQTableWidget()
        self.__datetime_since = QDateTimeEdit()
        self.__datetime_until = QDateTimeEdit()
        self.__button_query = QPushButton(self.__translate('', 'Query'))
        self.__button_export = QPushButton(self.__translate('', 'Export'))
        self.__button_clear = QPushButton(self.__translate('', 'Clear'))
        self.__check_identity_enable = QCheckBox(self.__translate('', 'Identity'))
        self.__check_datetime_enable = QCheckBox(self.__translate('', 'Datetime'))

        self.init_ui()

    # ---------------------------------------------------- UI Init -----------------------------------------------------

    def init_ui(self):
        self.__layout_control()
        self.__config_control()

    def __layout_control(self):
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        line = QHBoxLayout()
        line.addWidget(QLabel(self.__translate('', 'URI')), 0)
        line.addWidget(self.__combo_uri, 10)
        main_layout.addLayout(line)

        line = QHBoxLayout()
        line.addWidget(self.__check_identity_enable, 0)
        line.addWidget(self.__text_selected, 10)
        vline = QVBoxLayout()
        vline.addWidget(self.__button_reset)
        vline.addWidget(self.__button_select)
        line.addLayout(vline)
        main_layout.addLayout(line)

        line = QHBoxLayout()
        line.addWidget(self.__check_datetime_enable, 0)
        line.addWidget(self.__datetime_since, 10)
        line.addWidget(self.__datetime_until, 10)
        main_layout.addLayout(line)

        line = QHBoxLayout()
        line.addWidget(QLabel(' '), 10)
        line.addWidget(self.__button_query, 1)
        line.addWidget(self.__button_export, 1)
        line.addWidget(self.__button_clear, 1)
        main_layout.addLayout(line)

        main_layout.addWidget(self.__table_main)

    def __config_control(self):
        # self.__combo_uri.setEditable(True)
        self.__check_identity_enable.setChecked(True)
        self.__check_datetime_enable.setChecked(True)
        self.__datetime_since.setDateTime(default_since())
        self.__datetime_until.setDateTime(now())

        self.__button_select.clicked.connect(self.on_button_reset)
        self.__button_select.clicked.connect(self.on_button_select)
        self.__button_query.clicked.connect(self.on_button_query)
        self.__button_export.clicked.connect(self.on_button_export)
        self.__button_export.clicked.connect(self.on_button_clear)

        if self.__context is not None:
            data_agents = self.__context.get_sas_interface().sas_get_data_agent_probs()
            all_uri = [da.get('uri', '') for da in data_agents]
            for uri in all_uri:
                self.__combo_uri.addItem(uri)

        font = self.__text_selected.font()
        font_m = QFontMetrics(font)
        text_height = font_m.lineSpacing()
        self.__text_selected.setFixedHeight(3 * text_height)
        self.__text_selected.setEnabled(False)
        self.__update_selection_text()

        self.setMinimumSize(QSize(800, 600))
        self.setWindowTitle('Data Browser')

    def on_button_reset(self):
        self.__select_list.clear()
        self.__update_selection_text()

    def on_button_select(self):
        if self.__context is not None:
            dlg = SecuritiesPicker(self.__context.get_sas_interface())
            dlg.set_selection(self.__select_list)
            dlg.exec()
            if dlg.is_ok():
                self.__select_list = dlg.get_selection()
                self.__update_selection_text()

    def on_button_query(self):
        if len(self.__select_list) == 0:
            if QMessageBox().question(self, 'Query', 'Query for all securities may spends a lot of time.\nAre you sure?',
                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes) != QMessageBox.Yes:
                return
        self.__do_query()

    def on_button_export(self):
        if not self.__query_result.empty:
            dlg = ExportUi(self.__query_result)
            dlg.exec()
        else:
            QMessageBox().information(self, 'Warning', 'No data to export.')

    def on_button_clear(self):
        self.__query_result = pd.DataFrame()
        self.__show_result()

    # --------------------------------------------------------------------------------------

    def __do_query(self):
        uri = self.__combo_uri.currentText()
        # identity = self.__line_identity.text() if self.__check_identity_enable.isChecked() else None
        # identity = self.__combo_identity.get_input_securities() if self.__check_identity_enable.isChecked() else None

        identity = self.__select_list if (self.__check_identity_enable.isChecked() and
                                          len(self.__select_list) > 0) else None
        since = self.__datetime_since.dateTime().toPyDateTime() if self.__check_datetime_enable.isChecked() else None
        until = self.__datetime_until.dateTime().toPyDateTime() if self.__check_datetime_enable.isChecked() else None

        if self.__context is not None:
            result = self.__context.get_sas_interface().sas_query(uri, identity, (since, until))
        else:
            result = None

        if result is not None and not result.empty:
            self.__check_merge_result(result)
            self.__show_result(self.__query_result)
        else:
            QMessageBox().information(self, 'Query Result', 'Query data empty.')

    def __check_merge_result(self, result: pd.DataFrame):
        if result is None or result.empty:
            return
        if '_id' in result.columns:
            del result['_id']
        if self.__query_result is None or self.__query_result.empty:
            self.__query_result = result
            return
        intersect_columns = set(list(self.__query_result.columns) + list(result.columns))

        if len(intersect_columns) / len(result.columns) > 0.9:
            # Most columns is the same, suggest concat
            dlg = QueryOption('Most field matched, you can do concat.', 0)
        elif len(intersect_columns) >= 1:
            # Can be merged
            dlg = QueryOption('Has same fields, merge is valid.', 0)
        else:
            # Cannot concat nor merge, replace
            dlg = QueryOption('No match field. Only replace is available', 0, force_selection=True)
        dlg.exec()

        if not dlg.is_ok():
            return

        if dlg.get_selection() == 0:
            self.__query_result = result
        elif dlg.get_selection() == 1:
            self.__query_result = pd.concat([self.__query_result, result])
        else:
            # Not support
            assert False

        # TODO: How to generate an excel that multiple stock in one sheet with the trade date aligned?

    def __show_result(self, result: pd.DataFrame):
        write_df_to_qtable(result, self.__table_main)

    def __update_selection_text(self):
        if len(self.__select_list) > 0:
            self.__text_selected.setText(', '.join(self.__select_list))
        else:
            self.__text_selected.setText('<<All Securities>>')


# ----------------------------------------------------------------------------------------------------------------------

def main():
    app = QApplication(sys.argv)
    dlg = WrapperQDialog(DataHubUi(None))
    dlg.exec()


# ----------------------------------------------------------------------------------------------------------------------

def exception_hook(type, value, tback):
    # log the exception here
    print('Exception hook triggered.')
    print(type)
    print(value)
    print(tback)
    # then call the default handler
    sys.__excepthook__(type, value, tback)


if __name__ == "__main__":
    sys.excepthook = exception_hook
    try:
        main()
    except Exception as e:
        print('Error =>', e)
        print('Error =>', traceback.format_exc())
        exit()
    finally:
        pass
















